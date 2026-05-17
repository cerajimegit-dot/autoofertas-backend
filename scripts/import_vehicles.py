"""
Script para importar vehículos desde archivo Excel
Uso: python scripts/import_vehicles.py <file.xlsx> <enterprise_id> <branch_id>
"""

import os
import sys
import django
from pathlib import Path

# Configurar Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'playas_autos.settings')
django.setup()

import openpyxl
from decimal import Decimal
from core.models import Vehicle, Brand, VehicleModel, Branch, Enterprise, ExchangeRate


def import_vehicles(file_path, enterprise_id, branch_id):
    """Importar vehículos desde archivo Excel"""
    
    try:
        enterprise = Enterprise.objects.get(id=enterprise_id)
        branch = Branch.objects.get(id=branch_id, enterprise=enterprise)
    except (Enterprise.DoesNotExist, Branch.DoesNotExist) as e:
        print(f"❌ Error: Empresa o sucursal no encontrada - {e}")
        return
    
    # Cargar workbook
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    except Exception as e:
        print(f"❌ Error al abrir archivo: {e}")
        return
    
    success = 0
    errors = 0
    
    # Headers esperados (fila 1)
    headers = [cell.value for cell in ws[1]]
    
    print(f"\n📦 Importando vehículos desde: {file_path}")
    print(f"   Empresa: {enterprise.name}")
    print(f"   Sucursal: {branch.name}")
    print("-" * 80)
    
    # Iterar filas
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        try:
            values = [cell.value for cell in row]
            
            # Mapear valores
            marca = values[0]
            modelo = values[1]
            year = values[2]
            vin = values[3]
            placa = values[4]
            color = values[5]
            fob = Decimal(str(values[6]))
            conten = Decimal(str(values[7] or 0))
            despacho = Decimal(str(values[8] or 0))
            cam_vol = Decimal(str(values[9] or 0))
            price = Decimal(str(values[10]))
            moneda = values[11]
            cotizacion = values[12]
            estado = values[14] or 'available'
            notas = values[15] or ''
            
            # Validaciones
            if not all([marca, modelo, year, vin, fob, price]):
                raise ValueError("Falta datos requeridos (marca, modelo, año, VIN, FOB, precio)")
            
            # Obtener o crear marca
            brand = Brand.objects.filter(enterprise=enterprise, name=marca).first()
            if not brand:
                brand = Brand.objects.create(
                    enterprise=enterprise,
                    name=marca,
                    is_active=True
                )
                print(f"   ✓ Marca creada: {marca}")
            
            # Obtener o crear modelo
            model = VehicleModel.objects.filter(
                enterprise=enterprise,
                brand=brand,
                name=modelo
            ).first()
            if not model:
                model = VehicleModel.objects.create(
                    enterprise=enterprise,
                    brand=brand,
                    name=modelo,
                    is_active=True
                )
            
            # Validar que el VIN sea único
            if Vehicle.objects.filter(vin=vin).exists():
                raise ValueError(f"VIN duplicado: {vin}")
            
            # Obtener cotización si es USD
            exchange_rate = None
            if moneda == 'USD':
                if not cotizacion:
                    raise ValueError(f"Cotización requerida para USD (fila {row_idx})")
                exchange_rate = ExchangeRate.objects.filter(
                    enterprise=enterprise,
                    is_active=True
                ).first()
                
                if not exchange_rate:
                    raise ValueError(f"No hay cotización activa en la empresa")
            
            # Crear vehículo
            vehicle = Vehicle.objects.create(
                enterprise=enterprise,
                branch=branch,
                brand=brand,
                model=model,
                year=year,
                vin=vin,
                license_plate=placa or '',
                color=color or '',
                fob=fob,
                container=conten,
                dispatch=despacho,
                cam_vol=cam_vol,
                price=price,
                currency=moneda,
                exchange_rate=exchange_rate,
                state=estado,
                description=notas
            )
            
            print(f"   ✓ Vehículo creado: {brand.name} {model.name} ({year}) - VIN: {vin}")
            success += 1
            
        except Exception as e:
            errors += 1
            print(f"   ✗ Fila {row_idx}: {str(e)}")
    
    print("-" * 80)
    print(f"✅ Importación completada: {success} vehículos creados, {errors} errores")
    print()


if __name__ == '__main__':
    if len(sys.argv) < 4:
        print("Uso: python import_vehicles.py <archivo.xlsx> <enterprise_id> <branch_id>")
        print("Ejemplo: python import_vehicles.py sample_vehicles.xlsx 1 1")
        sys.exit(1)
    
    file_path = sys.argv[1]
    enterprise_id = int(sys.argv[2])
    branch_id = int(sys.argv[3])
    
    import_vehicles(file_path, enterprise_id, branch_id)
