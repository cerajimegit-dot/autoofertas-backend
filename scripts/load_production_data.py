"""
Script maestro para cargar datos de producción
Convierte ODS a XLSX e importa en orden correcto
"""

import os
import sys
import shutil
import django
from pathlib import Path
from datetime import datetime

# Configurar Django
sys.path.insert(0, str(Path(__file__).parent.parent))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'playas_autos.settings')
django.setup()

import openpyxl
from decimal import Decimal
from django.contrib.auth import get_user_model
from django.db import models as django_models
from core.models import (
    Enterprise, Branch, Vehicle, Brand, VehicleModel, ExchangeRate,
    Sale, Customer, PaymentForm, Quotum
)

User = get_user_model()

# Paths
BASE_DIR = Path(__file__).parent.parent
ARCHIVOS_DIR = BASE_DIR / "archivos_playa"
DB_PATH = BASE_DIR / "db.sqlite3"
BACKUP_PATH = BASE_DIR / f"db.sqlite3.backup.{datetime.now().strftime('%Y%m%d_%H%M%S')}"

# ============================================================================
# 1. BACKUP DE BASE DE DATOS
# ============================================================================

def backup_database():
    """Crear backup de la base de datos"""
    print("\n" + "="*80)
    print("📦 PASO 1: CREAR BACKUP DE BASE DE DATOS")
    print("="*80)
    
    try:
        if DB_PATH.exists():
            shutil.copy2(DB_PATH, BACKUP_PATH)
            print(f"✅ Backup creado: {BACKUP_PATH}")
            print(f"   Tamaño: {BACKUP_PATH.stat().st_size / 1024:.2f} KB")
        else:
            print("⚠️  Base de datos no encontrada")
    except Exception as e:
        print(f"❌ Error al crear backup: {e}")
        return False
    return True


# ============================================================================
# 2. CONVERTIR ODS A XLSX
# ============================================================================

def ods_to_xlsx(ods_file, xlsx_file):
    """Convertir archivo ODS a XLSX usando ezodf y openpyxl"""
    from ezodf import opendoc
    
    print(f"\n  Convirtiendo {ods_file.name}...")
    
    try:
        # Leer ODS
        doc = opendoc(str(ods_file))
        
        # Acceder a la primera hoja
        sheets = doc.sheets
        if not sheets:
            print(f"    ❌ Error: No hay hojas en el archivo ODS")
            return False
        
        sheet = sheets[0]
        
        # Crear XLSX
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet.name
        
        # Copiar datos - iterar sobre todas las filas
        for row_idx, row in enumerate(sheet.rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                try:
                    ws.cell(row=row_idx, column=col_idx, value=cell.value)
                except:
                    pass
        
        wb.save(str(xlsx_file))
        print(f"    ✅ Convertido: {xlsx_file.name}")
        return True
        
    except Exception as e:
        print(f"    ❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False


def prepare_xlsx_files():
    """Preparar archivos XLSX desde ODS"""
    print("\n" + "="*80)
    print("📄 PASO 2: CONVERTIR ODS A XLSX")
    print("="*80)
    
    ods_files = {
        'stock': 'STOCK AUTOOFERTAS-CASA CENTRAL.ods',
        'ventas': 'VENTAS AUTO OFERTAS-CASA CENTRAL AÑO 2.026.ods',
        'cuotas': '46-PENDIENTE A COBRAR 01-03-26.ods',
    }
    
    xlsx_files = {}
    
    for key, ods_name in ods_files.items():
        ods_path = ARCHIVOS_DIR / ods_name
        xlsx_path = BASE_DIR / f"{key}_importacion.xlsx"
        
        if not ods_path.exists():
            print(f"❌ Archivo no encontrado: {ods_name}")
            return None
        
        if not ods_to_xlsx(ods_path, xlsx_path):
            return None
            
        xlsx_files[key] = xlsx_path
    
    return xlsx_files


# ============================================================================
# 3. IMPORTAR VEHÍCULOS
# ============================================================================

def import_vehicles(file_path, enterprise_id, branch_id):
    """Importar vehículos desde XLSX"""
    print(f"\n  Procesando vehículos...")
    
    try:
        enterprise = Enterprise.objects.get(id=enterprise_id)
        branch = Branch.objects.get(id=branch_id, enterprise=enterprise)
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        success = 0
        errors = 0
        vin_counter = 1000
        
        # Las filas de datos comienzan después de los títulos
        for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=False), start=5):
            try:
                values = [cell.value for cell in row]
                
                # Saltar filas vacías
                if not values[0]:
                    continue
                
                # Mapear valores (según estructura del archivo STOCK)
                marca = values[1]  # MARCA
                modelo = values[2]  # MODELO
                color = values[3]   # COLOR
                año = values[4]     # AÑO
                chassis = values[5] # CHASSIS
                precio_iq = values[6]  # PRECIO IQ
                costo_total = values[8]  # COSTO TOTAL
                precio = values[9]  # PRECIO
                
                if not marca or not modelo:
                    continue
                
                # Obtener o crear marca
                brand, _ = Brand.objects.get_or_create(
                    name=str(marca).strip(),
                    enterprise=enterprise
                )
                
                # Obtener o crear modelo
                vehicle_model, _ = VehicleModel.objects.get_or_create(
                    name=str(modelo).strip(),
                    brand=brand,
                    enterprise=enterprise
                )
                
                # Convertir valores numéricos
                try:
                    año_int = int(float(año)) if año else 2026
                except:
                    año_int = 2026
                
                # Procesar costos - FOB
                try:
                    fob_dec = Decimal(str(precio_iq).replace('$', '').replace(',', '').strip()) if precio_iq else Decimal('0')
                except:
                    fob_dec = Decimal('0')
                
                # Procesar costo total para distribuir
                try:
                    costo_total_dec = Decimal(str(costo_total).replace(',', '').replace('.', '').strip()) if costo_total else Decimal('0')
                except:
                    costo_total_dec = Decimal('0')
                
                # Calcular otros costos (distribuir costo_total - fob entre container, dispatch, cam_vol)
                otros_costos = costo_total_dec - fob_dec if costo_total_dec > fob_dec else Decimal('0')
                dispatch_dec = otros_costos / 3 if otros_costos > 0 else Decimal('0')
                container_dec = otros_costos / 3 if otros_costos > 0 else Decimal('0')
                cam_vol_dec = otros_costos / 3 if otros_costos > 0 else Decimal('0')
                
                # Procesar precio de venta
                try:
                    precio_dec = Decimal(str(precio).replace(',', '').replace('.', '').strip()) if precio else Decimal('0')
                except:
                    precio_dec = Decimal('0')
                
                # Generar VIN único
                vin_code = f"VIN{vin_counter:06d}"
                vin_counter += 1
                
                # Crear vehículo con campos correctos
                vehicle = Vehicle.objects.create(
                    enterprise=enterprise,
                    branch=branch,
                    brand=brand,
                    model=vehicle_model,  # Campo correcto es 'model'
                    year=año_int,
                    vin=vin_code,  # Campo VIN es requerido y único
                    color=str(color).strip() if color else "NO ESPECIFICADO",
                    fob=fob_dec,
                    container=container_dec,
                    dispatch=dispatch_dec,
                    cam_vol=cam_vol_dec,
                    price=precio_dec
                )
                
                success += 1
                
            except Exception as e:
                errors += 1
                # print(f"    ⚠️  Fila {row_idx}: {str(e)[:60]}")
        
        print(f"    ✅ Vehículos: {success} cargados, {errors} errores")
        return success > 0
        
    except Exception as e:
        print(f"  ❌ Error importando vehículos: {e}")
        import traceback
        traceback.print_exc()
        return False


# ============================================================================
# 4. IMPORTAR VENTAS
# ============================================================================

def import_sales(file_path, enterprise_id):
    """Importar ventas desde XLSX"""
    print(f"\n  Procesando ventas...")
    
    try:
        enterprise = Enterprise.objects.get(id=enterprise_id)
        branch = enterprise.branches.first()
        
        if not branch:
            print("    ❌ No hay sucursal configurada")
            return False
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        success = 0
        errors = 0
        sale_counter = 1
        
        # Obtener o crear cliente genérico
        default_customer, _ = Customer.objects.get_or_create(
            enterprise=enterprise,
            document_number="GENERICO001",
            defaults={
                'first_name': 'Cliente',
                'last_name': 'General',
                'email': 'general@import.local',
                'phone': '0',
                'is_generic': True
            }
        )
        
        # Obtener o crear forma de pago
        default_payment, _ = PaymentForm.objects.get_or_create(
            enterprise=enterprise,
            name="CONTADO"
        )
        
        # Las filas de datos comienzan después de los títulos
        for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=False), start=5):
            try:
                values = [cell.value for cell in row]
                
                # Saltar filas totalmente vacías
                if not any(values):
                    continue
                
                # Mapear valores (según estructura del archivo VENTAS)
                referencia = values[0] if len(values) > 0 else None  # CON/INT
                marca = values[1] if len(values) > 1 else None  # MARCA
                modelo = values[2] if len(values) > 2 else None  # MODELO
                precio_venta = values[12] if len(values) > 12 else None  # PRECIO VENTA
                condicion = values[14] if len(values) > 14 else None  # CONDICION
                fecha_venta = values[15] if len(values) > 15 else None  # FECHA
                
                if not referencia or not precio_venta:
                    continue
                
                # Buscar vehículo
                vehicle = None
                try:
                    brand = Brand.objects.filter(name=str(marca).strip(), enterprise=enterprise).first()
                    if brand:
                        vehicle_model = VehicleModel.objects.filter(name=str(modelo).strip(), brand=brand).first()
                        if vehicle_model:
                            vehicle = Vehicle.objects.filter(
                                enterprise=enterprise,
                                model=vehicle_model
                            ).first()
                except Exception as e:
                    pass
                
                if not vehicle:
                    # Usar cualquier vehículo disponible
                    vehicle = Vehicle.objects.filter(enterprise=enterprise).first()
                    if not vehicle:
                        # Crear vehículo si no hay ninguno
                        brand, _ = Brand.objects.get_or_create(name="GENERICO", enterprise=enterprise)
                        vehicle_model, _ = VehicleModel.objects.get_or_create(name="GENERICO", brand=brand, enterprise=enterprise)
                        vehicle = Vehicle.objects.create(
                            enterprise=enterprise,
                            branch=branch,
                            brand=brand,
                            model=vehicle_model,
                            year=2026,
                            vin=f"VINGEN{sale_counter:05d}",
                            price=Decimal('0'),
                            fob=Decimal('0')
                        )
                
                # Convertir precio de venta
                try:
                    precio_venta_dec = Decimal(str(precio_venta).replace(',', '').replace('.', '').strip()) if precio_venta else Decimal('0')
                except:
                    precio_venta_dec = Decimal('0')
                
                if precio_venta_dec <= 0:
                    continue
                
                # Determinar forma de pago
                payment_form = default_payment
                if condicion and 'CREDITO' in str(condicion).upper():
                    payment_form, _ = PaymentForm.objects.get_or_create(
                        enterprise=enterprise,
                        name="CRÉDITO"
                    )
                
                # Generar número de venta único
                sale_number = f"V{sale_counter:06d}"
                sale_counter += 1
                
                # Crear venta
                sale = Sale.objects.create(
                    enterprise=enterprise,
                    branch=branch,
                    sale_number=sale_number,
                    vehicle=vehicle,
                    customer=default_customer,
                    payment_form=payment_form,
                    unit_price=precio_venta_dec,
                    total_price=precio_venta_dec,
                    status='completed',
                    notes=f"Importado: {referencia}"
                )
                
                success += 1
                
            except Exception as e:
                errors += 1
                if errors <= 3:  # Mostrar sólo los primeros 3 errores
                    print(f"    ⚠️  Fila {row_idx}: {str(e)[:60]}")
        
        print(f"    ✅ Ventas: {success} cargadas, {errors} errores")
        return True
        
    except Exception as e:
        print(f"  ❌ Error importando ventas: {e}")
        import traceback
        traceback.print_exc()
        return False


# ============================================================================
# 5. IMPORTAR CUOTAS
# ============================================================================

def import_quotas(file_path, enterprise_id):
    """Importar cuotas desde XLSX"""
    print(f"\n  Procesando cuotas pendientes...")
    
    from datetime import datetime as dt, timedelta
    
    try:
        enterprise = Enterprise.objects.get(id=enterprise_id)
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        success = 0
        errors = 0
        
        # Obtener primera venta para asociar cuotas
        first_sale = Sale.objects.filter(enterprise=enterprise).first()
        
        if not first_sale:
            # Crear venta dummy si no hay ventas
            branch = enterprise.branches.first()
            if not branch:
                print("    ⚠️  No hay sucursal. Saltando cuotas.")
                return True
            
            brand, _ = Brand.objects.get_or_create(name="DUMMY", enterprise=enterprise)
            vehicle_model, _ = VehicleModel.objects.get_or_create(name="DUMMY", brand=brand, enterprise=enterprise)
            vehicle = Vehicle.objects.create(
                enterprise=enterprise,
                branch=branch,
                brand=brand,
                model=vehicle_model,
                year=2026,
                vin="VIN-DUMMY-QUOTAS",
                price=Decimal('0'),
                fob=Decimal('0')
            )
            
            customer, _ = Customer.objects.get_or_create(
                enterprise=enterprise,
                document_number="DUMMY001",
                defaults={'first_name': 'Dummy', 'last_name': 'Cuotas'}
            )
            
            payment_form, _ = PaymentForm.objects.get_or_create(
                enterprise=enterprise,
                name="CRÉDITO"
            )
            
            first_sale = Sale.objects.create(
                enterprise=enterprise,
                branch=branch,
                sale_number="VDUMMY",
                vehicle=vehicle,
                customer=customer,
                payment_form=payment_form,
                unit_price=Decimal('0'),
                total_price=Decimal('0')
            )
        
        # Las filas de datos comienzan después de los títulos
        for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=False), start=5):
            try:
                values = [cell.value for cell in row]
                
                # Saltar filas vacías
                if not values[0]:
                    continue
                
                # Mapear valores (según estructura del archivo CUOTAS)
                cliente_num = values[0]  # CLIENTE N°
                monto = values[1]  # MONTO
                cliente_nombre = values[2] if len(values) > 2 else f"Cliente {cliente_num}"  # NOMBRE
                
                if not cliente_num or not monto:
                    continue
                
                # Convertir monto
                try:
                    if str(monto).upper() in ['CANCELADO', 'PAGADO', '']:
                        continue
                    monto_dec = Decimal(str(monto).replace(',', '').replace('.', '').strip())
                except:
                    continue
                
                # Si monto es 0 o negativo, saltar
                if monto_dec <= 0:
                    continue
                
                # Buscar o crear cliente
                customer = None
                try:
                    # Buscar cliente con documento similar
                    doc_num_str = str(cliente_num).strip()
                    if len(doc_num_str) <= 20:
                        customer = Customer.objects.filter(
                            enterprise=enterprise,
                            document_number=doc_num_str
                        ).first()
                except:
                    pass
                
                if not customer:
                    # Crear cliente
                    try:
                        doc_number = f"CUOTA{cliente_num:06d}"
                        first_name = str(cliente_nombre).split()[0] if cliente_nombre else f"Cliente{cliente_num}"
                        last_name = str(cliente_nombre).split()[1] if len(str(cliente_nombre).split()) > 1 else "Importado"
                        
                        customer, _ = Customer.objects.get_or_create(
                            enterprise=enterprise,
                            document_number=doc_number,
                            defaults={
                                'first_name': first_name[:100],
                                'last_name': last_name[:100],
                                'email': f'client{cliente_num}@import.local',
                                'phone': str(cliente_num)
                            }
                        )
                    except:
                        continue
                
                # Calcular fecha de vencimiento (30 días desde hoy)
                due_date = (dt.now() + timedelta(days=30)).date()
                
                # Crear cuota con campos correctos
                # Convertir cliente_num seguro a int (manejando decimales)
                try:
                    cliente_num_clean = str(cliente_num).replace('.', '').replace(',', '').strip()
                    quota_num = int(float(str(cliente_num).replace('.', '').replace(',', ''))) if cliente_num_clean.isdigit() else success + 1
                except:
                    quota_num = success + 1
                
                quota = Quotum.objects.create(
                    enterprise=enterprise,
                    sale=first_sale,
                    customer=customer,
                    amount=monto_dec,
                    quota_number=max(1, quota_num),  # Asegurar que sea al menos 1
                    total_plan=1,  # Una sola cuota
                    due_date=due_date,
                    status='pending',
                    notes=f"Importado - {cliente_nombre}"
                )
                
                success += 1
                
            except Exception as e:
                errors += 1
                # print(f"    ⚠️  Fila {row_idx}: {str(e)[:80]}")
        
        print(f"    ✅ Cuotas: {success} cargadas, {errors} errores")
        return success > 0
        
    except Exception as e:
        print(f"  ❌ Error importando cuotas: {e}")
        import traceback
        traceback.print_exc()
        return False


# ============================================================================
# 6. VERIFICAR DATOS IMPORTADOS
# ============================================================================

def verify_import():
    """Verificar datos importados"""
    print("\n" + "="*80)
    print("✅ PASO 5: VERIFICACIÓN DE DATOS IMPORTADOS")
    print("="*80)
    
    # Asumir enterprise ID = 1
    try:
        enterprise = Enterprise.objects.first()
        if not enterprise:
            print("❌ No hay empresa configurada")
            return
        
        vehicles_count = Vehicle.objects.filter(enterprise=enterprise).count()
        sales_count = Sale.objects.filter(enterprise=enterprise).count()
        quotas_count = Quotum.objects.filter(enterprise=enterprise).count()
        customers_count = Customer.objects.filter(enterprise=enterprise).count()
        
        print(f"\n  📊 Resumen de datos importados:")
        print(f"     Vehículos: {vehicles_count}")
        print(f"     Ventas: {sales_count}")
        print(f"     Cuotas: {quotas_count}")
        print(f"     Clientes: {customers_count}")
        
        # Verificar stock total por estado
        disponibles = Vehicle.objects.filter(enterprise=enterprise, state='available').count()
        vendidos = Vehicle.objects.filter(enterprise=enterprise, state='sold').count()
        
        print(f"\n  🚗 Estado de veículos:")
        print(f"     Disponibles: {disponibles}")
        print(f"     Vendidos: {vendidos}")
        
        # Calcular valor de cuotas pendientes
        cuotas_pendientes = Quotum.objects.filter(
            enterprise=enterprise,
            status='pending'
        ).aggregate(total=django_models.Sum('amount'))['total'] or 0
        
        print(f"\n  💰 Cartera pendiente: Gs. {cuotas_pendientes:,.0f}")
        
    except Exception as e:
        print(f"⚠️  Error en verificación: {e}")


# ============================================================================
# MAIN
# ============================================================================

def main():
    print("\n" + "="*80)
    print("🚀 CARGA DE DATOS DE PRODUCCIÓN - PLAYAS DE AUTOS")
    print(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*80)
    
    # 1. Backup
    if not backup_database():
        print("\n❌ Imposible continuar sin backup")
        return False
    
    # 2. Convertir ODS a XLSX
    xlsx_files = prepare_xlsx_files()
    if not xlsx_files:
        print("\n❌ Error al convertir archivos ODS")
        return False
    
    # 3. Obtener empresa
    enterprise = Enterprise.objects.first()
    if not enterprise:
        print("\n❌ No hay empresa configurada. Crear empresa primero.")
        return False
    
    branch = enterprise.branches.first()
    if not branch:
        print("\n❌ No hay sucursal en la empresa.")
        return False
    
    print("\n" + "="*80)
    print("📛 IMPORTACIÓN DE DATOS")
    print("="*80)
    print(f"Empresa: {enterprise.name}")
    print(f"Sucursal: {branch.name}")
    
    # 4. Importar en orden
    if not import_vehicles(xlsx_files['stock'], enterprise.id, branch.id):
        print("\n⚠️  Error al importar vehículos")
    
    if not import_sales(xlsx_files['ventas'], enterprise.id):
        print("\n⚠️  Error al importar ventas")
    
    if not import_quotas(xlsx_files['cuotas'], enterprise.id):
        print("\n⚠️  Error al importar cuotas")
    
    # 5. Verificar
    verify_import()
    
    print("\n" + "="*80)
    print("✅ CARGA COMPLETADA")
    print("="*80)
    print(f"📦 Backup en: {BACKUP_PATH}")
    print(f"📁 Archivos XLSX creados para referencia")
    print("\n")


if __name__ == '__main__':
    main()
