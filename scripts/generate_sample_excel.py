"""
Script para generar archivos de prueba en Excel con datos de ejemplo
para pruebas de importación masiva.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import os


def create_vehicles_template():
    """Crear archivo de plantilla de vehículos para importación"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehículos"
    
    # Headers
    headers = [
        'Marca', 'Modelo', 'Año', 'VIN', 'Placa', 'Color',
        'FOB', 'CONTEN', 'DESPACHO', 'CAM/VOL', 'Precio', 'Moneda',
        'Cotización (si es USD)', 'Sucursal', 'Estado', 'Notas'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de ejemplo
    data = [
        ['Toyota', 'Corolla', 2020, 'ABC123456', 'ABC-123', 'Blanco', 15000, 500, 300, 100, 18900, 'USD', 7850, 'Sucursal 1', 'Disponible', ''],
        ['Honda', 'Civic', 2019, 'DEF789012', 'DEF-456', 'Negro', 16000, 500, 300, 100, 20000, 'USD', 7850, 'Sucursal 1', 'Disponible', ''],
        ['Ford', 'F-150', 2021, 'GHI345678', 'GHI-789', 'Gris', 20000, 600, 400, 150, 25200, 'USD', 7850, 'Sucursal 2', 'Disponible', ''],
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save('sample_vehicles.xlsx')
    print("✓ Archivo 'sample_vehicles.xlsx' creado")


def create_customers_template():
    """Crear archivo de plantilla de clientes para importación"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Headers
    headers = [
        'Nombre', 'Apellido', 'Tipo Doc', 'Número Doc', 'Email',
        'Teléfono', 'Dirección', 'Ciudad', 'Notas'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de ejemplo
    data = [
        ['Juan', 'García', 'CI', '1234567', 'juan@email.com', '+595971111111', 'Calle 1', 'Asunción', ''],
        ['María', 'López', 'CI', '7654321', 'maria@email.com', '+595971112222', 'Calle 2', 'Asunción', ''],
        ['Carlos', 'Martínez', 'RUC', '80000001', 'carlos@empresa.com', '+595972223333', 'Calle 3', 'Encarnación', ''],
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Ajustar ancho
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save('sample_customers.xlsx')
    print("✓ Archivo 'sample_customers.xlsx' creado")


def create_sales_template():
    """Crear archivo de plantilla de ventas para importación"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"
    
    # Headers
    headers = [
        'Número Venta', 'Fecha Venta', 'Cliente', 'Doc Cliente', 'VIN Vehículo',
        'Precio Unitario', 'Descuento', 'Precio Total', 'Forma Pago', 'Vendedor', 'Notas'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de ejemplo
    today = datetime.now().strftime('%Y-%m-%d')
    data = [
        ['V001', today, 'Juan García', '1234567', 'ABC123456', 18900, 500, 18400, 'Efectivo', 'vendor1', ''],
        ['V002', today, 'María López', '7654321', 'DEF789012', 20000, 0, 20000, 'Tarjeta', 'vendor1', ''],
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Ajustar ancho
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save('sample_sales.xlsx')
    print("✓ Archivo 'sample_sales.xlsx' creado")


def create_quotas_template():
    """Crear archivo de plantilla de cuotas para importación"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuotas"
    
    # Headers
    headers = [
        'Número Venta', 'Cliente', 'Doc Cliente', 'Número Cuota', 'Plan',
        'Total Cuotas', 'Monto Cuota', 'Interés', 'Fecha Vencimiento', 'Notas'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de ejemplo
    today = datetime.now()
    data = [
        ['V001', 'Juan García', '1234567', 1, '6 Cuotas', 6, 3066.67, 33.33, (today + timedelta(days=30)).strftime('%Y-%m-%d'), ''],
        ['V001', 'Juan García', '1234567', 2, '6 Cuotas', 6, 3066.67, 33.33, (today + timedelta(days=60)).strftime('%Y-%m-%d'), ''],
        ['V001', 'Juan García', '1234567', 3, '6 Cuotas', 6, 3066.66, 33.34, (today + timedelta(days=90)).strftime('%Y-%m-%d'), ''],
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Ajustar ancho
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save('sample_quotas.xlsx')
    print("✓ Archivo 'sample_quotas.xlsx' creado")


if __name__ == '__main__':
    print("Generando archivos de prueba en Excel...")
    print()
    
    create_vehicles_template()
    create_customers_template()
    create_sales_template()
    create_quotas_template()
    
    print()
    print("¡Archivos de prueba generados exitosamente!")
    print("Archivos creados:")
    print("  - sample_vehicles.xlsx")
    print("  - sample_customers.xlsx")
    print("  - sample_sales.xlsx")
    print("  - sample_quotas.xlsx")
